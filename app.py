from flask import Flask, render_template, request, redirect, url_for, session, flash, make_response, send_file, jsonify
from models import db, User, Exam, Question, Result, Log
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from weasyprint import HTML
import os
import csv
import json
import openai
from dotenv import load_dotenv

# 환경 변수 로드
load_dotenv()

app = Flask(__name__)
app.secret_key = "your_secret_key"
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///C:/Users/JS/Downloads/JS/instance/exam.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# OpenAI API 설정
openai.api_key = os.getenv('OPENAI_API_KEY')

db.init_app(app)

# 데이터베이스 초기화 및 관리자 계정 생성
with app.app_context():
    db.create_all()
    # 관리자 계정이 없으면 생성
    admin = User.query.filter_by(username='admin').first()
    if not admin:
        admin = User(username='admin', nickname='관리자', password='j12209942!', role='admin')
        db.session.add(admin)
        db.session.commit()
        print("관리자 계정이 생성되었습니다. (아이디: admin, 비밀번호: j12209942!)")
    else:
        # 기존 관리자 비밀번호 업데이트
        admin.password = 'j12209942!'
        db.session.commit()
        print("관리자 비밀번호가 업데이트되었습니다. (아이디: admin, 비밀번호: j12209942!)")

@app.route('/')
def home():
    return render_template('home.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        nickname = request.form['nickname']
        password = request.form['password']
        role = request.form['role']
        user = User(username=username, nickname=nickname, password=password, role=role)
        db.session.add(user)
        db.session.commit()
        flash('회원가입 완료')
        return redirect(url_for('login'))
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username, password=password).first()
        if user:
            session['user_id'] = user.id
            session['role'] = user.role
            return redirect(url_for('exam_form', exam_id=1)) if user.role == 'student' else redirect(url_for('admin_dashboard'))
        else:
            flash('로그인 실패')
    return render_template('login.html')

@app.route('/exam_form/<int:exam_id>', methods=['GET', 'POST'])
def exam_form(exam_id):
    user_id = session.get('user_id')
    if request.method == 'POST':
        answers = request.form
        questions = Question.query.filter_by(exam_id=exam_id).all()
        earned_score = 0
        total_score = 0
        answer_dict = {}
        for q in questions:
            selected = answers.get(str(q.id))
            answer_dict[q.id] = int(selected) if selected else None
            # 배점이 설정되어 있지 않으면 기본값 1로 설정
            question_score = q.score if q.score is not None else 1
            total_score += question_score
            if selected and int(selected) == q.answer:
                earned_score += question_score
        # 총점이 0이면 기본 100점으로 계산
        if total_score == 0:
            total_score = len(questions)
            earned_score = sum(1 for q in questions if answer_dict.get(q.id) == q.answer)
        score = round((earned_score / total_score) * 100, 1) if total_score > 0 else 0
        result = Result(student_id=user_id, exam_id=exam_id, score=score, answers=json.dumps(answer_dict))
        db.session.add(result)
        db.session.commit()
        return redirect(url_for('grade_result', result_id=result.id))
    questions = Question.query.filter_by(exam_id=exam_id).all()
    return render_template('exam_form.html', questions=questions)

@app.route('/grade_result/<int:result_id>')
def grade_result(result_id):
    result = db.session.get(Result, result_id)
    student = db.session.get(User, result.student_id)
    exam = db.session.get(Exam, result.exam_id)
    questions = Question.query.filter_by(exam_id=exam.id).order_by(Question.question_number).all()
    student_answers = json.loads(result.answers) if result.answers else {}
    question_results = []
    for q in questions:
        student_answer = student_answers.get(str(q.id)) if isinstance(student_answers, dict) else student_answers.get(q.id)
        is_correct = (student_answer == q.answer)
        question_results.append({
            'number': q.question_number,
            'student_answer': student_answer,
            'correct_answer': q.answer,
            'is_correct': is_correct
        })
    return render_template('grade_result.html', result=result, student=student, exam=exam, question_results=question_results)

@app.route('/admin_dashboard')
def admin_dashboard():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    exams = Exam.query.all()
    exam_labels = []
    exam_averages = []
    exam_maxes = []
    for exam in exams:
        results = Result.query.filter_by(exam_id=exam.id).all()
        if results:
            scores = [r.score for r in results]
            exam_labels.append(exam.title)
            exam_averages.append(sum(scores) / len(scores))
            exam_maxes.append(max(scores))
    return render_template('admin_dashboard.html',
                           exam_labels=exam_labels,
                           exam_averages=exam_averages,
                           exam_maxes=exam_maxes)

@app.route('/upload_exam', methods=['GET', 'POST'])
def upload_exam():
    if request.method == 'POST':
        file = request.files['file']
        exam_id = request.form.get('exam_id')
        if not file or not exam_id:
            flash("모든 정보를 입력해주세요.")
            return redirect(request.url)
        filename = secure_filename(file.filename)
        filepath = os.path.join('uploads', filename)
        file.save(filepath)
        wb = load_workbook(filepath)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 번호, 보기1, 보기2, 보기3, 보기4, 보기5, 정답
            if len(row) < 7:
                continue  # 필수 열이 부족하면 건너뜀
            _, c1, c2, c3, c4, c5, ans = row
            answer_num = None
            # 정답이 숫자(1~5)면 그대로, 텍스트면 일치하는 보기 번호로 변환
            if ans is not None:
                try:
                    answer_num = int(ans)
                    if answer_num < 1 or answer_num > 5:
                        answer_num = None
                except:
                    # 텍스트로 입력된 경우
                    choices = [c1, c2, c3, c4, c5]
                    for idx, choice in enumerate(choices, 1):
                        if choice and str(choice).strip() == str(ans).strip():
                            answer_num = idx
                            break
            question = Question(
                exam_id=exam_id,
                choice1=c1, choice2=c2, choice3=c3,
                choice4=c4, choice5=c5, answer=answer_num
            )
            db.session.add(question)
        db.session.commit()
        # 자동배점 적용
        recalculate_scores(exam_id)
        flash("시험 업로드가 완료되었습니다.")
        return redirect(url_for('admin_dashboard'))
    return render_template('upload_exam.html')

@app.route('/my_stats')
def my_stats():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user = db.session.get(User, session['user_id'])
    results = Result.query.filter_by(user_id=user.id).order_by(Result.timestamp).all()
    labels = [f"시험 {r.exam_id}" for r in results]
    scores = [r.score for r in results]
    return render_template('my_stats.html', labels=labels, scores=scores, user=user)

@app.route('/export_stats')
def export_stats():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))

    results = Result.query.all()
    output = []
    output.append(['시험 제목', '학생 ID', '닉네임', '반', '점수', '응시 시간'])

    for r in results:
        user = db.session.get(User, r.user_id)
        exam = db.session.get(Exam, r.exam_id)
        output.append([
            exam.title,
            user.username,
            user.nickname,
            user.class_name or '',
            r.score,
            r.timestamp.strftime('%Y-%m-%d %H:%M')
        ])

    si = ""
    for row in output:
        si += ",".join([str(cell) for cell in row]) + "\n"

    response = make_response(si)
    response.headers["Content-Disposition"] = "attachment; filename=exam_results.csv"
    response.headers["Content-type"] = "text/csv"
    return response

def renumber_questions(exam_id):
    questions = Question.query.filter_by(exam_id=exam_id).order_by(Question.question_number).all()
    for idx, q in enumerate(questions, 1):
        q.question_number = idx
    db.session.commit()

@app.route('/edit_exam/<int:exam_id>', methods=['GET', 'POST'])
def edit_exam(exam_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    
    exam = Exam.query.get_or_404(exam_id)
    questions = Question.query.filter_by(exam_id=exam_id).order_by(Question.question_number).all()
    
    if request.method == 'POST':
        print("=== 시험 편집 폼 제출 ===")
        
        # 기존 문항 수
        existing_count = len(questions)
        
        # 폼에서 전송된 문항 수 확인 (숫자 입력란들 개수)
        submitted_questions = []
        i = 1
        while f'number_{i}' in request.form:
            submitted_questions.append(i)
            i += 1
        
        submitted_count = len(submitted_questions)
        print(f"기존 문항 수: {existing_count}, 제출된 문항 수: {submitted_count}")
        
        # 문항 수가 변경된 경우 처리
        if submitted_count > existing_count:
            # 새 문항 추가
            for i in range(existing_count + 1, submitted_count + 1):
                new_question = Question(
                    exam_id=exam_id,
                    question_number=i,
                    choice1='보기 1',
                    choice2='보기 2', 
                    choice3='보기 3',
                    choice4='보기 4',
                    choice5='보기 5',
                    answer='',
                    score=0
                )
                db.session.add(new_question)
                print(f"새 문항 추가: {i}번")
        
        elif submitted_count < existing_count:
            # 문항 삭제 (마지막 문항부터 삭제)
            questions_to_delete = questions[submitted_count:]
            for question in questions_to_delete:
                db.session.delete(question)
                print(f"문항 삭제: {question.question_number}번")
        
        # 문항 데이터 업데이트
        total_score = 0
        for i in submitted_questions:
            question_text = request.form.get(f'question_{i}', '')
            answer = request.form.get(f'answer_{i}', '')
            score_str = request.form.get(f'score_{i}', '0')
            score = float(score_str) if score_str.strip() else 0.0
            
            # 기존 문항 찾기 또는 새 문항 생성
            if i <= existing_count:
                question = questions[i-1]
            else:
                question = Question.query.filter_by(exam_id=exam_id, question_number=i).first()
            
            if question:
                # 기존 문항 업데이트
                if question.answer != answer:
                    print(f"문항 {question.question_number} 정답 변경: {question.answer} -> {answer}")
                if question.score != score:
                    print(f"문항 {question.question_number} 배점 변경: {question.score} -> {score}")
                
                question.choice1 = request.form.get(f'choice1_{i}', '')
                question.choice2 = request.form.get(f'choice2_{i}', '')
                question.choice3 = request.form.get(f'choice3_{i}', '')
                question.choice4 = request.form.get(f'choice4_{i}', '')
                question.choice5 = request.form.get(f'choice5_{i}', '')
                question.answer = answer
                question.score = score
                total_score += score
        
        # 총점 업데이트
        exam.total_score = total_score
        print(f"총점 자동 계산: {total_score}점")
        
        try:
            db.session.commit()
            flash('시험이 성공적으로 저장되었습니다.', 'success')
            print("시험 편집 저장 완료")
        except Exception as e:
            db.session.rollback()
            print(f"시험 편집 저장 오류: {e}")
            flash('시험 저장 중 오류가 발생했습니다.', 'error')
        
        return redirect(url_for('edit_exam', exam_id=exam_id))
    
    # GET 요청: 템플릿 렌더링
    # 문항 데이터를 템플릿에 맞게 가공
    questions_data = []
    for q in questions:
        choices = [
            {'number': 1, 'text': q.choice1},
            {'number': 2, 'text': q.choice2},
            {'number': 3, 'text': q.choice3},
            {'number': 4, 'text': q.choice4},
            {'number': 5, 'text': q.choice5}
        ]
        questions_data.append({
            'id': q.id,
            'number': q.question_number,
            'text': f"문항 {q.question_number}",
            'choices': choices,
            'answer': q.answer,
            'score': q.score
        })
    
    return render_template(
        'edit_exam.html',
        exam=exam,
        questions=questions_data
    )

@app.route('/logs')
def view_logs():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    logs = Log.query.order_by(Log.timestamp.desc()).all()
    return render_template('logs.html', logs=logs)

@app.route('/copy_exam_ui/<int:exam_id>', methods=['GET', 'POST'])
def copy_exam_ui(exam_id):
    original = Exam.query.get_or_404(exam_id)
    if request.method == 'POST':
        new_title = request.form['new_title']
        copied = Exam(title=new_title, created_by=original.created_by)
        db.session.add(copied)
        db.session.commit()

        questions = Question.query.filter_by(exam_id=exam_id).all()
        for idx, q in enumerate(questions):
            new_q = Question(
                exam_id=copied.id,
                choice1=q.choice1,
                choice2=q.choice2,
                choice3=q.choice3,
                choice4=q.choice4,
                choice5=q.choice5,
                answer=q.answer,
                score=q.score,
                question_number=idx + 1
            )
            db.session.add(new_q)

        log = Log(action=f"Exam copied: {original.title} -> {new_title}", user_id=session.get('user_id'))
        db.session.add(log)
        db.session.commit()
        return redirect(url_for('admin_dashboard'))
    return render_template('copy_exam.html', exam=original)

@app.route('/admin/students', methods=['GET', 'POST'])
def admin_students():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    # 학생 추가
    if request.method == 'POST':
        username = request.form['username']
        nickname = request.form['nickname']
        password = request.form['password']
        grade = request.form['grade']
        class_part = request.form['class_name']
        class_name = f"{grade}-{class_part}"
        user = User(username=username, nickname=nickname, password=password, role='student', class_name=class_name)
        db.session.add(user)
        db.session.commit()
        flash('학생이 추가되었습니다.')
        return redirect(url_for('admin_students'))
    # 반별 필터
    grade_filter = request.args.get('grade_filter', '')
    class_filter = request.args.get('class_filter', '')
    filter_query = User.query.filter_by(role='student')
    if grade_filter:
        if class_filter:
            filter_query = filter_query.filter(User.class_name == f"{grade_filter}-{class_filter}")
        else:
            filter_query = filter_query.filter(User.class_name.like(f"{grade_filter}-%"))
    elif class_filter:
        filter_query = filter_query.filter(User.class_name.like(f"%-{class_filter}"))
    students = filter_query.all()
    # 학년-반 목록 생성 (중복 없이)
    class_names = db.session.query(User.class_name).filter(User.role=='student').distinct().all()
    class_names = [c[0] for c in class_names if c[0]]
    return render_template('admin_students.html', students=students, class_names=class_names, grade_filter=grade_filter, class_filter=class_filter)

@app.route('/admin/students/edit/<int:user_id>', methods=['GET', 'POST'])
def edit_student(user_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    user = User.query.get_or_404(user_id)
    if request.method == 'POST':
        user.username = request.form['username']
        user.nickname = request.form['nickname']
        user.password = request.form['password']
        grade = request.form['grade']
        class_part = request.form['class_name']
        user.class_name = f"{grade}-{class_part}"
        db.session.commit()
        flash('학생 정보가 수정되었습니다.')
        return redirect(url_for('admin_students'))
    # 학년/반 분리해서 템플릿에 전달
    grade = ''
    class_part = ''
    if user.class_name and '-' in user.class_name:
        grade, class_part = user.class_name.split('-', 1)
    return render_template('edit_student.html', user=user, grade=grade, class_part=class_part)

@app.route('/admin/students/delete/<int:user_id>', methods=['POST'])
def delete_student(user_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    user = User.query.get_or_404(user_id)
    db.session.delete(user)
    db.session.commit()
    flash('학생이 삭제되었습니다.')
    return redirect(url_for('admin_students'))

@app.route('/admin/class_stats')
def class_stats():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    # 반별로 학생과 성적 집계
    class_data = []
    class_names = db.session.query(User.class_name).filter(User.role=='student').distinct().all()
    class_names = [c[0] for c in class_names if c[0]]
    for cname in class_names:
        students = User.query.filter_by(role='student', class_name=cname).all()
        student_ids = [s.id for s in students]
        results = Result.query.filter(Result.student_id.in_(student_ids)).all()
        scores = [r.score for r in results]
        avg_score = round(sum(scores)/len(scores), 2) if scores else 0
        max_score = max(scores) if scores else 0
        class_data.append({
            'class_name': cname,
            'student_count': len(students),
            'avg_score': avg_score,
            'max_score': max_score
        })
    return render_template('class_stats.html', class_data=class_data)

@app.route('/admin/students/<int:user_id>/results')
def student_results(user_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    user = User.query.get_or_404(user_id)
    results = Result.query.filter_by(student_id=user_id).order_by(Result.timestamp.desc()).all()
    exams = {e.id: e for e in Exam.query.all()}
    return render_template('student_results.html', user=user, results=results, exams=exams)

@app.route('/create_exam', methods=['GET', 'POST'])
def create_exam():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    if request.method == 'POST':
        title = request.form['title']
        category = request.form.get('category', '기타')
        question_count = int(request.form.get('question_count', 5))
        scoring_method = request.form.get('scoring_method', 'auto')
        
        exam = Exam(title=title, created_by=session['user_id'], category=category)
        db.session.add(exam)
        db.session.commit()
        
        # 문항 수만큼 빈 문제 생성
        for i in range(1, question_count+1):
            q = Question(
                exam_id=exam.id, 
                question_number=i,
                choice1='보기 1', 
                choice2='보기 2', 
                choice3='보기 3', 
                choice4='보기 4', 
                choice5='보기 5',
                answer='',
                score=0
            )
            db.session.add(q)
        
        # 배점 설정 방식에 따른 배점 계산
        if scoring_method == 'manual':
            # 수동 설정: 기본 배점과 총점 사용
            default_score_str = request.form.get('default_score', '0')
            total_score_str = request.form.get('total_score', '0')
            
            # 빈 문자열 체크 후 안전한 형변환
            default_score = float(default_score_str) if default_score_str.strip() else 0.0
            total_score = float(total_score_str) if total_score_str.strip() else 0.0
            
            if default_score > 0:
                # 기본 배점으로 모든 문항 설정
                for q in Question.query.filter_by(exam_id=exam.id).all():
                    q.score = round(default_score, 1)
                exam.total_score = total_score
                print(f"수동 배점 설정 완료: 문항 {question_count}개, 문항당 {default_score}점, 총점 {total_score}점")
            else:
                # 기본 배점이 설정되지 않은 경우 기본값 사용
                default_score = 20.0
                total_score = question_count * default_score
                for q in Question.query.filter_by(exam_id=exam.id).all():
                    q.score = round(default_score, 1)
                exam.total_score = total_score
                print(f"기본값으로 배점 설정: 문항 {question_count}개, 문항당 {default_score}점, 총점 {total_score}점")
        else:
            # 자동 균등 분배: 사용자가 설정한 총점 사용
            auto_total_score_str = request.form.get('auto_total_score', '100')
            total_score = float(auto_total_score_str) if auto_total_score_str.strip() else 100.0
            score_per_question = total_score / question_count
            for q in Question.query.filter_by(exam_id=exam.id).all():
                q.score = round(score_per_question, 1)
            exam.total_score = total_score
            print(f"자동 균등 분배 완료: 문항 {question_count}개, 문항당 {score_per_question:.1f}점, 총점 {total_score}점")
        
        db.session.commit()
        
        return redirect(url_for('edit_exam', exam_id=exam.id))
    return render_template('create_exam.html')

@app.route('/download_sample_exam')
def download_sample_exam():
    return send_file('sample_exam.xlsx', as_attachment=True)

@app.route('/api/update_question_number', methods=['POST'])
def update_question_number():
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'error': '권한이 없습니다.'}), 403
    
    data = request.get_json()
    qid = data.get('qid')  # 실제 질문 ID
    new_number = data.get('new_number')
    exam_id = data.get('exam_id')
    
    print(f"번호 변경 요청: qid={qid}, new_number={new_number}, exam_id={exam_id}")
    
    if not qid or not new_number or not exam_id:
        return jsonify({'error': '필수 파라미터가 누락되었습니다.'}), 400
    
    # 실제 질문 ID로 질문 찾기
    question = db.session.get(Question, qid)
    
    if question and question.exam_id == int(exam_id):
        old_number = question.question_number
        target_number = int(new_number)
        
        # 해당 시험의 모든 질문 가져오기
        all_questions = Question.query.filter_by(exam_id=exam_id).order_by(Question.question_number).all()
        total_questions = len(all_questions)
        
        print(f"총 문항 수: {total_questions}, 요청된 번호: {target_number}")
        
        # 번호 유효성 검사
        if target_number < 1 or target_number > total_questions:
            return jsonify({'error': f'번호는 1부터 {total_questions}까지 입력 가능합니다.'}), 400
        
        # 번호가 실제로 변경되었는지 확인
        if old_number == target_number:
            return jsonify({'success': True, 'message': '번호가 변경되지 않았습니다.'})
        
        # 새로운 번호 재정렬 로직
        try:
            # 1. 모든 문항을 임시로 9999로 설정 (중복 방지)
            for q in all_questions:
                q.question_number = 9999
            
            # 2. 대상 문항을 원하는 위치에 배치
            question.question_number = target_number
            
            # 3. 나머지 문항들을 순서대로 배치
            current_number = 1
            for q in all_questions:
                if q.id != question.id:  # 대상 문항 제외
                    # 변경된 번호보다 작은 위치에 배치
                    if current_number < target_number:
                        q.question_number = current_number
                        current_number += 1
                    else:
                        # 변경된 번호 다음부터 배치
                        q.question_number = current_number + 1
                        current_number += 1
            
            db.session.commit()
            
            print(f"번호 변경 성공: 문항 {question.id} 번호 {old_number} -> {target_number}")
            print("전체 번호 재정렬 완료")
            
            # 업데이트된 번호 목록 반환
            updated_questions = Question.query.filter_by(exam_id=exam_id).order_by(Question.question_number).all()
            updated_numbers = [{'qid': q.id, 'new_number': q.question_number} for q in updated_questions]
            
            return jsonify({
                'success': True, 
                'message': f'번호가 {old_number}에서 {target_number}로 변경되었습니다.',
                'updated_numbers': updated_numbers
            })
            
        except Exception as e:
            db.session.rollback()
            print(f"번호 변경 중 오류 발생: {str(e)}")
            return jsonify({'error': f'번호 변경 중 오류가 발생했습니다: {str(e)}'}), 500
    
    return jsonify({'error': '질문을 찾을 수 없습니다.'}), 404

# ChatGPT API 관련 라우트들
@app.route('/chatgpt/chat', methods=['GET', 'POST'])
def chatgpt_chat():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        user_message = request.form.get('message', '')
        
        if not user_message:
            flash('메시지를 입력해주세요.')
            return render_template('chatgpt_chat.html')
        
        try:
            # ChatGPT API 호출
            response = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": "당신은 친절하고 도움이 되는 AI 어시스턴트입니다. 한국어로 답변해주세요."},
                    {"role": "user", "content": user_message}
                ],
                max_tokens=1000,
                temperature=0.7
            )
            
            ai_response = response.choices[0].message.content
            
            # 로그 저장
            log = Log(
                user_id=session['user_id'],
                action=f"ChatGPT 사용: {user_message[:50]}...",
                details=f"사용자: {user_message}\nAI: {ai_response}"
            )
            db.session.add(log)
            db.session.commit()
            
            return render_template('chatgpt_chat.html', 
                                 user_message=user_message, 
                                 ai_response=ai_response)
        
        except Exception as e:
            flash(f'ChatGPT API 호출 중 오류가 발생했습니다: {str(e)}')
            return render_template('chatgpt_chat.html')
    
    return render_template('chatgpt_chat.html')

@app.route('/chatgpt/generate_questions', methods=['GET', 'POST'])
def generate_questions():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        topic = request.form.get('topic', '')
        num_questions = int(request.form.get('num_questions', 5))
        difficulty = request.form.get('difficulty', '보통')
        
        if not topic:
            flash('주제를 입력해주세요.')
            return render_template('generate_questions.html')
        
        try:
            # ChatGPT로 문제 생성 요청
            prompt = f"""
            다음 조건에 맞는 객관식 문제 {num_questions}개를 생성해주세요:
            - 주제: {topic}
            - 난이도: {difficulty}
            - 형식: 번호, 보기1, 보기2, 보기3, 보기4, 보기5, 정답번호
            - JSON 형식으로 응답해주세요
            
            예시 형식:
            [
                {{
                    "question_number": 1,
                    "choice1": "보기1",
                    "choice2": "보기2", 
                    "choice3": "보기3",
                    "choice4": "보기4",
                    "choice5": "보기5",
                    "answer": 3
                }}
            ]
            """
            
            response = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": "당신은 교육 전문가입니다. 객관식 문제를 생성해주세요."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=2000,
                temperature=0.7
            )
            
            ai_response = response.choices[0].message.content
            
            # JSON 파싱 시도
            try:
                import json
                questions_data = json.loads(ai_response)
                return render_template('generate_questions.html', 
                                     topic=topic,
                                     questions_data=questions_data,
                                     ai_response=ai_response)
            except json.JSONDecodeError:
                return render_template('generate_questions.html', 
                                     topic=topic,
                                     ai_response=ai_response)
        
        except Exception as e:
            flash(f'문제 생성 중 오류가 발생했습니다: {str(e)}')
            return render_template('generate_questions.html')
    
    return render_template('generate_questions.html')

@app.route('/chatgpt/analyze_results', methods=['GET', 'POST'])
def analyze_results():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        exam_id = request.form.get('exam_id')
        
        if not exam_id:
            flash('시험을 선택해주세요.')
            return render_template('analyze_results.html')
        
        # 시험 결과 데이터 수집
        results = Result.query.filter_by(exam_id=exam_id).all()
        exam = db.session.get(Exam, exam_id)
        
        if not results:
            flash('해당 시험의 결과가 없습니다.')
            return render_template('analyze_results.html')
        
        # 결과 분석을 위한 데이터 준비
        scores = [r.score for r in results]
        avg_score = sum(scores) / len(scores)
        max_score = max(scores)
        min_score = min(scores)
        
        analysis_data = {
            'exam_title': exam.title,
            'total_students': len(results),
            'average_score': round(avg_score, 2),
            'max_score': max_score,
            'min_score': min_score,
            'score_distribution': {
                '90-100': len([s for s in scores if s >= 90]),
                '80-89': len([s for s in scores if 80 <= s < 90]),
                '70-79': len([s for s in scores if 70 <= s < 80]),
                '60-69': len([s for s in scores if 60 <= s < 70]),
                '0-59': len([s for s in scores if s < 60])
            }
        }
        
        try:
            # ChatGPT로 결과 분석 요청
            prompt = f"""
            다음 시험 결과를 분석하고 개선 방안을 제시해주세요:
            
            시험명: {analysis_data['exam_title']}
            응시자 수: {analysis_data['total_students']}명
            평균 점수: {analysis_data['average_score']}점
            최고 점수: {analysis_data['max_score']}점
            최저 점수: {analysis_data['min_score']}점
            
            점수 분포:
            - 90-100점: {analysis_data['score_distribution']['90-100']}명
            - 80-89점: {analysis_data['score_distribution']['80-89']}명
            - 70-79점: {analysis_data['score_distribution']['70-79']}명
            - 60-69점: {analysis_data['score_distribution']['60-69']}명
            - 0-59점: {analysis_data['score_distribution']['0-59']}명
            
            이 결과를 바탕으로 교육적 개선 방안을 제시해주세요.
            """
            
            response = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": "당신은 교육 평가 전문가입니다. 시험 결과를 분석하고 개선 방안을 제시해주세요."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=1500,
                temperature=0.7
            )
            
            ai_analysis = response.choices[0].message.content
            
            return render_template('analyze_results.html', 
                                 analysis_data=analysis_data,
                                 ai_analysis=ai_analysis)
        
        except Exception as e:
            flash(f'결과 분석 중 오류가 발생했습니다: {str(e)}')
            return render_template('analyze_results.html')
    
    # GET 요청 시 시험 목록 제공
    exams = Exam.query.all()
    return render_template('analyze_results.html', exams=exams)

# ✅ DB 생성 + 서버 실행
if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True, port=5001)
